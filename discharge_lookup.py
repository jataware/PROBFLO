

import openpyxl
import pandas as pd
from functools import cache
from netica import NeticaGraph



@cache
def max_contiguous_alignment_score(A:str, B:str) -> int:
    """
    Determine the maximum amount of overlap between two strings
    
    Used to determine what site the blocks map to in the discharge ranges spreadsheet
    """
    if len(A) > len(B):
        A, B = B, A
    get_score = lambda A,B: sum(1 for a,b in zip(A, B) if a == b)
    scores = [get_score(A, B[shift:]) for shift in range(len(B) - len(A) + 1)]
    return max(scores)

@cache
def get_varname_map(varnames:tuple[str,...], sites:tuple[str,...], strip=True) -> dict[str, str]:
    """
    map from site name to the variable name used by the discharge ranges spreadsheet
    """
    var_name_map = {}
    for var in varnames:
        #compare the variable name to each site
        if strip:
            scores = [max_contiguous_alignment_score(site.strip(), var.strip()) for site in sites]
        else:
            scores = [max_contiguous_alignment_score(site, var) for site in sites]

        #determine which site matched best
        i, score = max(enumerate(scores), key=lambda x: x[1])
        best_site = sites[i]

        match_percent = score / len(best_site.strip() if strip else best_site)

        assert match_percent > 0.8, f"No good match for '{var}' in possible sites: {sites}"

        if match_percent < 1:
            print(f"Warning: '{var}' is only a {match_percent:.0%} match for site '{best_site}'")

        # print(f"'{var}' -> '{best_site}'")
        var_name_map[best_site] = var

    return var_name_map



@cache
def get_discharge_scenario_data():

    # load the official list of site names
    file_map_df = pd.read_csv('neta/limpopo_27_subbasin/risk_region_mapping.csv')
    sites: list[str] = file_map_df['Site'].tolist()


    # Load the discharge scenario workbook and select the worksheet
    file_path = "neta/limpopo_27_subbasin/2023-04-03_Discharge_ranges_for_WM.xlsx"
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook["MEDIAN RANKS "] #TODO: sheet name has a space at the end...

    # Define the row indices for each scenario
    NATURAL_ROW = 3
    PRESENT_ROW = 4
    E_FLOW_ROW = 5
    FUTURE_ROW = 6
    # scenarios = ['NATURAL', 'PRESENT', 'E-FLOW', 'FUTURE']
    variables = ['DISCHARGE_YR', 'DISCHARGE_LF', 'DISCHARGE_HF', 'DISCHARGE_FD']

    # Define starting column index (B is column index 2)
    starting_column = 2

    # Define the column ranges for the data blocks
    data_blocks = [(starting_column, starting_column + 3), (starting_column + 5, starting_column + 8), (starting_column + 10, starting_column + 13)]
    last_column = starting_column + 133 #last column based on how much data is expected in the sheet

    print('Collecting Discharge Scenario Data...')

    # Initialize the variables
    subbasins = []
    discharge_values = {
        "NATURAL": {},
        "PRESENT": {},
        "E-FLOW": {},
        "FUTURE": {},
    }

    # Loop through the columns and extract the titles and values
    for col_start in range(starting_column, last_column, 5):
        col_end = col_start + 3
        data_blocks.append((col_start, col_end))

        # Extract the titles (merged cells)
        target_cell_address = sheet.cell(row=1, column=col_start).coordinate
        for merged_cell_range in sheet.merged_cells.ranges:
            if target_cell_address in merged_cell_range:
                top_left_cell_address = merged_cell_range.coord.split(':')[0]
                subbasin = sheet[top_left_cell_address].value
                subbasins.append(subbasin)
                break

        # Extract the values for each scenario's row
        discharge_values['NATURAL'][subbasin] = dict(zip(variables, tuple(sheet.cell(row=NATURAL_ROW, column=col).value for col in range(col_start, col_end + 1))))
        discharge_values['PRESENT'][subbasin] = dict(zip(variables, tuple(sheet.cell(row=PRESENT_ROW, column=col).value for col in range(col_start, col_end + 1))))
        discharge_values['E-FLOW'][subbasin]  = dict(zip(variables, tuple(sheet.cell(row=E_FLOW_ROW, column=col).value for col in range(col_start, col_end + 1))))
        discharge_values['FUTURE'][subbasin]  = dict(zip(variables, tuple(sheet.cell(row=FUTURE_ROW, column=col).value for col in range(col_start, col_end + 1))))


    # Create a map from site name to the variable name used by the discharge ranges spreadsheet
    varname_map = get_varname_map(tuple(subbasins), tuple(sites), strip=False)


    return varname_map, discharge_values




def update_net_discharge_scenario(site:str, net:NeticaGraph, scenario:str):
    '''
    Update all net variables for the given discharge scenario at the particular site specified
    '''
    varname_map, discharge_values = get_discharge_scenario_data()

    assert site in varname_map, f"Site '{site}' not found in varname_map"
    assert scenario in discharge_values, f"Value '{scenario}' not a valid discharge scenario. Expected one of {', '.join(discharge_values.keys())}"

    for node, value in discharge_values[scenario][varname_map[site]].items():
        net.enter_finding(node, value, retract=True, verbose=True)
